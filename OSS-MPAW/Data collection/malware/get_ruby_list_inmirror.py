import subprocess
import pandas as pd
import openpyxl

def download_packages(excel_file, sheet_name, start_row, column_index, ruby_registry_mirrors):
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook[sheet_name]

        row_index = start_row
        cell_value1 = sheet.cell(row=row_index, column=column_index).value
        cell_value2 = sheet.cell(row=row_index, column=column_index + 1).value

        while cell_value1:
            package_name = str(cell_value1).strip()  # Get the value of the cell and strip whitespace
            if cell_value2:
                package_version = str(cell_value2).strip() 
                for mirror in ruby_registry_mirrors:
                    download_command = f"gem fetch {package_name} -v {package_version} -s {mirror}"   
                    print(f"Downloading package: {package_name}")
                    try:
                        subprocess.run(download_command, shell=True, check=True)
                        # If downloaded successfully, input "get" into the cell in the second column on the same row
                        success_cell = sheet.cell(row=row_index, column=column_index + 2)
                        success_cell.value = "get"
                        break
                    except subprocess.CalledProcessError:
                        # If download fails, input "untraced" into the cell in the second column on the same row
                        failure_cell = sheet.cell(row=row_index, column=column_index + 2)
                        failure_cell.value = "untraced"
            else:
                for mirror in ruby_registry_mirrors:
                    download_command = f"gem fetch {package_name} -s {mirror}"   
                    print(f"Downloading package: {package_name}")
                    try:
                        subprocess.run(download_command, shell=True, check=True)
                        # If downloaded successfully, input "get" into the cell in the second column on the same row
                        success_cell = sheet.cell(row=row_index, column=column_index + 2)
                        success_cell.value = "get"
                        break
                    except subprocess.CalledProcessError:
                        # If download fails, input "untraced" into the cell in the second column on the same row
                        failure_cell = sheet.cell(row=row_index, column=column_index + 2)
                        failure_cell.value = "untraced"

            row_index += 1
            cell_value1 = sheet.cell(row=row_index, column=column_index).value
            cell_value2 = sheet.cell(row=row_index, column=column_index + 1).value

        workbook.save(excel_file)  
        print('Download completed.')
    except Exception as e:
        print('Error:', str(e))

ruby_registry_mirrors = ['https://gems.ruby-china.com','http://ruby.taobao.org',
           'https://mirrors.tuna.tsinghua.edu.cn/rubygems/','https://mirrors.hust.edu.cn/rubygems/',
           'https://mirrors.aliyun.com/rubygems/','http://mirror.sysu.edu.cn/rubygems/',
           'http://ruby.sdutlinux.org/']

excel_file_path = 'Get_Malicious_Package/malicious_packages.xlsx'
sheet_name = 'rubygems'
start_row_index = 2
column_index = 2

download_packages(excel_file_path, sheet_name, start_row_index, column_index, ruby_registry_mirrors)
